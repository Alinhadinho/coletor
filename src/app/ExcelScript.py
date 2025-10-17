import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import Font, PatternFill
from datetime import datetime
import os

# ... (seu dicionário de produtos de exemplo pode ser removido ou mantido para testes)

def encontrar_primeira_linha_vazia(worksheet):
    # ... (esta função permanece igual)
    max_row = worksheet.max_row if worksheet.max_row > 1 else 1
    for row_num in range(max_row, 0, -1):
        cell_value = worksheet.cell(row=row_num, column=1).value
        if cell_value is not None and str(cell_value).strip() != "":
            return row_num + 1
    return 1

# MODIFICADO: A função agora aceita um diretório de saída
def adicionar_e_formatar_produtos(nome_arquivo_original: str, produtos: list, nome_planilha: str = 'Planilha1', output_dir: str = '.'):
    
    if not os.path.exists(nome_arquivo_original):
        print(f"❌ ERRO: O arquivo de template '{nome_arquivo_original}' não foi encontrado.")
        return None
        
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    nome_base, extensao = os.path.splitext(os.path.basename(nome_arquivo_original))
    # MODIFICADO: Nome do arquivo de saída
    nome_arquivo_copia = f"{nome_base}_EXPORTADO_{timestamp}{extensao}"
    caminho_completo_copia = os.path.join(output_dir, nome_arquivo_copia)
    
    try:
        workbook = openpyxl.load_workbook(nome_arquivo_original)
    except Exception as e:
        print(f"❌ Erro ao carregar o arquivo: {e}")
        return None

    # ... (O restante da lógica de manipulação da planilha permanece igual)
    if nome_planilha in workbook.sheetnames:
        worksheet = workbook[nome_planilha]
    else:
        worksheet = workbook.active
    
    proxima_linha_livre = encontrar_primeira_linha_vazia(worksheet)
    
    # MODIFICADO: A verificação de cabeçalho agora verifica a coluna A na linha 2
    if proxima_linha_livre <= 2 and worksheet.cell(row=2, column=1).value is None:
        # Se a linha 2 (onde os dados deveriam começar) estiver vazia, assume-se que os cabeçalhos são necessários
        # headers = ["Produto", "EAN", "PLU", "Quantidade", "Validade"]
        # for col_num, header in enumerate(headers, 1):
        #      worksheet.cell(row=proxima_linha_livre, column=col_num, value=header)
        # proxima_linha_livre += 1
        pass # O template já tem cabeçalho

    row_index_start = proxima_linha_livre
    current_row = row_index_start
    
    for produto in produtos:
        data_validade = datetime.strptime(produto.get("validade", "01/01/1900"), "%d/%m/%Y").date()
        
        # Ordem das colunas igual ao template
        linha_dados = [
            produto.get("EAN", ""),
            produto.get("nome", ""),
            data_validade,
            produto.get("quantidade", 0),
        ]
        
        for col_num, valor in enumerate(linha_dados, 1):
            worksheet.cell(row=current_row, column=col_num, value=valor)
            
        current_row += 1
    
    last_inserted_row = current_row - 1
    
    # Formatação (se houver alguma)
    for row_num in range(row_index_start, last_inserted_row + 1):
        worksheet.cell(row=row_num, column=3).number_format = 'DD/MM/YYYY' # Coluna da Data
        worksheet.cell(row=row_num, column=4).number_format = '0'        # Coluna da Quantidade

    try:
        workbook.save(caminho_completo_copia)
        print(f"\n✅ SUCESSO! Dados exportados para: \n'{caminho_completo_copia}'")
        return caminho_completo_copia # Retorna o caminho onde o arquivo foi salvo
    except PermissionError:
        print(f"\n❌ ERRO: Permissão negada. Verifique se o arquivo está aberto.")
        return None